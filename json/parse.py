#!/usr/bin/python
import json
import openpyxl
import time
from datetime import datetime
from tkinter import *
from tkinter import Tk, Label, Button, Entry, StringVar, W, E, filedialog
from tkinter.ttk import *


# Input Token and Choose File
class File_Token_Chooser:

    def __init__(self, master):
        self.master = master
        master.title("Keyword parser")

        self.token = ""
        self.filepath = StringVar()


        self.labelT = Label(master, text="Token:")
        self.entry1 = Entry(master, textvariable=self.token) #, validate="key", validatecommand=(vcmd, '%P'))
        self.entry2 = Entry(master)
        self.labelF = Label(master, text="File Path:")

        self.start_cell = Entry(master, text=1)
        self.labelStart = Label(master, text='Номер первой ячейки:')

        self.end_cell = Entry(master, text=2)
        self.labelEnd = Label(master, text='Номер последней ячейки:')

        self.ok_button = Button(master, text="OK", command=master.quit)
        self.choose_button = Button(master, text="Choose file", command=lambda: self.choose())

        # LAYOUT

        self.entry1.grid(row=1, column=1, columnspan=3, sticky=W+E)
        self.labelT.grid(row=1, column=0)
        self.entry2.grid(row=2, column=1, columnspan=3, sticky=W + E)
        self.labelF.grid(row=2, column=0)

        self.start_cell.grid(row=3, column=1, columnspan=1, sticky=W + E)
        self.labelStart.grid(row=3, column=0)
        self.end_cell.grid(row=4, column=1, columnspan=1, sticky=W + E)
        self.labelEnd.grid(row=4, column=0)

        self.ok_button.grid(row=5, column=0)
        self.choose_button.grid(row=2, column=5)

    def get_filepath(self):

        return self.filepath

    def get_toke(self):
        return self.token

    def choose(self):
        self.filepath = filedialog.askopenfilename()

# Running GUI input
root = Tk()
my_gui = File_Token_Chooser(root)
root.mainloop()

# Getting Data from GUI input
token = my_gui.get_toke()
filepath = my_gui.get_filepath()
start_cell = my_gui.start_cell.get()
end_cell = my_gui.end_cell.get()



# Output to Excel Class
def output_to_excel():
    now = datetime.now()
    output_file = 'results_' + now.strftime("%Y-%m-%d_%H:%M:%S") + '.xlsx'
    wb = openpyxl.Workbook()
    engine = "g_us"
    sheetname = "Top_keyword" + "_" + engine
    sheet = wb.active
    sheet.title = sheetname
    # wb.create_sheet(sheetname)
    # sheet = wb.get_sheet_by_name(sheetname)

    sheet.cell(row=1 + 1, column=1).value = 'Keyword'
    sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=1, column=2).value = 'Query engine'
    sheet.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=1, column=3).value = 'Response Message'
    sheet.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=1, column=4).value = 'Queries count'
    sheet.cell(row=1, column=4).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=1, column=5).value = 'Found Results'
    sheet.cell(row=1, column=5).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=1, column=6).value = 'Domains'
    sheet.cell(row=1, column=6).font = openpyxl.styles.Font(bold=True)
    sheet.cell(row=1, column=7).value = 'Full URLs'
    sheet.cell(row=1, column=7).font = openpyxl.styles.Font(bold=True)
    # ws = wb.activate()

    wb.save(output_file)

# Open example JSON file
file = open("json.example","r")
# Parse example JSON file
jsondata = json.loads(file.read())


# Test. Outputting to Console data from JSON
domains = ""
for item in jsondata['result']['top']:
    domains += item.get("domain") + ","
urls = ""
for item in jsondata['result']['top']:
    urls += item.get("url") + ","
print(jsondata['left_lines'])
print(jsondata['result']['results'])
print("found_domains:" + domains)
print("found_urls:" + urls)
print(len(jsondata['result']['top']))
if jsondata['status_code'] == 200:
    print("status code is 200:")
    print(jsondata['status_code'])

print(jsondata['status_msg'])
status_msg = "%i"% (jsondata['status_code']) + ", " + jsondata['status_msg']
print(status_msg)


# Write Data to Excel File
output_to_excel()

root = Tk()
T = Text(root, height=2, width=30)
T.pack()
T.insert(END, "Done")
mainloop()
