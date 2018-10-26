#!/usr/bin/python
import json
import urllib.request as urlrequest
from urllib.parse import urlencode
from openpyxl import load_workbook
from datetime import datetime
import openpyxl
from tkinter import *
from tkinter import filedialog

host = 'http://api.serpstat.com/v3'
method = 'keyword_top'
# se = 'g_ru'


# Input Token and Choose File
class FileTokenChooser:

    def __init__(self, master):
        self.master = master
        master.title("Keyword parser")

        self.token = ""
        self.filepath = StringVar()
        self.greeter = Text(master)
        self.greeter.insert(INSERT, "Добро пожаловать в парсер запросов по ключевым словам"
                                    "\nДля парсинга введите ваши данные:"
                                    "\nТокен, Выберите файл Excel .xlsx. Ключевые слова нужно указывать в первом столбце."
                                    "\nТакже нужно ввести номера первой и последней ячеек, содержащих ключевые слова для поиска"
                                    "\n\nТакже нужно указать поисковый движок (g_su, g_en и тд)."
                                    "\nПолный список есть тут https://serpstat.com/api/6-request-parameters/ "
                                    "\n\nПосле чего нажмите кнопку ОК"
                                    "\nПо завершении программа выведет сообщение с именем файла")

        # self.lst1 = []
        self.se = StringVar(master)
        self.se.set("g_ru")
        self.w = OptionMenu(master, self.se, "g_us","g_uk",
                            "g_au","g_ru","g_ca","g_bg","g_ua",
                            "g_za","g_lt","g_lv","g_by","g_kz",
                            "g_it","g_es","g_fr","g_de","g_nl",
                            "g_br","g_il","g_dk","y_213","y_2",
                            "y_187","y_54")

        self.labelT = Label(master, text="Token:")
        self.entry1 = Entry(master, textvariable=self.token)
        self.entry2 = Entry(master)
        self.labelF = Label(master, text="File Path:")

        self.start_cell = Entry(master, text=1)
        self.labelStart = Label(master, text='Номер первой ячейки:')

        self.end_cell = Entry(master, text=2)
        self.labelEnd = Label(master, text='Номер последней ячейки:')

        self.ok_button = Button(master, text="OK", command=lambda: self.run_program())
        self.quit_button = Button(master, text='Quit', command=master.destroy)
        self.choose_button = Button(master, text="Choose file", command=lambda: self.choose())

        # LAYOUT

        self.entry1.grid(row=1, column=1, columnspan=3, sticky=W+E)
        self.labelT.grid(row=1, column=0)
        self.entry2.grid(row=2, column=1, columnspan=3, sticky=W + E)
        self.labelF.grid(row=2, column=0)

        self.start_cell.grid(row=3, column=1, columnspan=1, sticky=W + E)
        self.labelStart.grid(row=3, column=0)
        self.end_cell.grid(row=4, column=1, columnspan=1, sticky=W + E)
        self.labelEnd.grid(row=4, column=0)

        self.ok_button.grid(row=5, column=0)
        self.quit_button.grid(row=5, column=1)
        self.choose_button.grid(row=2, column=5)

        self.greeter.grid(row=6, column=0)
        self.w.grid(row=6, column=1)

    def run_program(self):
        # Get variables from input window
        self.newWindow = Toplevel(self.master)
        self.app = MsgWindow(self.newWindow)
        self.app.display_msg1("В процессе выполнения возникла ошибка: ", "test case")

        my_token = self.get_toke()
        file_path = self.get_filepath()
        start_cell = self.start_cell.get()
        end_cell = self.end_cell.get()
        se = self.get_se()

        # Opening Excel Workbook with input queries
        wb1 = load_workbook(file_path)
        sheet_input = wb1.get_sheet_by_name('Лист1')

        # Preparing new output file
        now = datetime.now()
        output_file = 'results_' + now.strftime("%Y-%m-%d_%H:%M:%S") + '.xlsx'
        wb2 = openpyxl.Workbook()
        sheetname = method + "_" + se
        wb2.create_sheet(sheetname)
        sheet = wb2.get_sheet_by_name('Top_keyword')

        # Preparing output column names
        sheet.cell(row=1, column=1).value = 'Keyword'
        sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=1, column=2).value = 'Response Message'
        sheet.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=1, column=3).value = 'Found Results'
        sheet.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=1, column=4).value = 'Domains'
        sheet.cell(row=1, column=4).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=1, column=5).value = 'Full URLs'
        sheet.cell(row=1, column=5).font = openpyxl.styles.Font(bold=True)

        # Sending queries to web API and getting data from responses
        for i in range(int(start_cell), int(end_cell)):
            qr = sheet_input.cell(row=i, column=1).value.replace(" ", "%20")
            params = {
                'query': qr,  # string for get info
                'se': se,  # string search engine
                'token': my_token,  # string personal token
            }
            api_url = "{host}/{method}?{params}".format(
                host=host,
                method=method,
                params=urlencode(params)
            )
            try:
                json_data = urlrequest.urlopen(api_url).read()
            except Exception as e0:
                print("API request error: {error}".format(error=e0))
                self.newWindow = Toplevel(self.master)
                self.app = MsgWindow(self.newWindow)
                self.app.display_msg(e0)


            # Parsing JSON
            jsondata = json.loads(json_data)

            # Extracting data from JSON
            domains = ""
            urls = ""
            left_lines = jsondata['left_lines']
            status_msg = "%i" % (jsondata['status_code']) + ", " + jsondata['status_msg']
            results_found = ""

            if jsondata['status_code'] == 200:
                results_found = jsondata['result']['results']
                # Getting domains
                for item in jsondata['result']['top']:
                    domains += item.get("domain") + ","
                # Getting URLs
                for item in jsondata['result']['top']:
                    urls += item.get("url") + ","

            print(qr)
            print(status_msg)
            print(results_found)
            print(domains)
            print(urls)
            print("lines left: " + results_found)
            sheet.cell(row=i + 1, column=1).value = qr.replace("%20", " ")
            sheet.cell(row=i + 1, column=2).value = status_msg
            sheet.cell(row=i + 1, column=3).value = results_found
            sheet.cell(row=i + 1, column=4).value = domains
            sheet.cell(row=i + 1, column=5).value = urls

        # Writing output to file
        wb2.save(output_file)

        # Opening window with results
        root2 = Tk()
        T = Text(root2, height=4, width=30)
        B1 = Button(text="Ok", command=root2.destroy)
        T.pack()
        B1.pack()

        T.insert(END, "Done")
        T.insert(END, "\nОсталось запросов по токену:" + left_lines)
        T.insert(END, "\nРезультат сохранен в файл: " + output_file)

    def get_se(self):
        return self.se.get()

    def get_filepath(self):
        return self.filepath

    def get_toke(self):
        return self.token

    def choose(self):
        self.filepath = filedialog.askopenfilename()


class MsgWindow:
    def __init__(self, master):
        self.master = master
        self.frame = Frame(self.master)
        self.quitButton = Button(self.frame, text = 'Quit', width = 25, command = self.close_windows)
        self.TF = Text(self.master, height=4, width=30)
        self.TF.pack()
        self.quitButton.pack()
        self.frame.pack()
        self.TF.insert(END, "В процессе выполнения возникла ошибка")

    def display_msg1(self,msg1, error_msg):
        self.TF.insert(END, "\n" + msg1 + error_msg)

    def display_msg2(self, msg2):
        self.TF.insert(END, "\n" + msg2)

    def close_windows(self):
        self.master.destroy()


# Running Main Window
def main():
    root = Tk()
    my_gui = FileTokenChooser(root)
    root.mainloop()

if __name__ == '__main__':
    main()


"""
#  Getting Data from GUI input
my_token = my_gui.get_toke()
file_path = my_gui.get_filepath()
start_cell = my_gui.start_cell.get()
end_cell = my_gui.end_cell.get()
se = my_gui.get_se()

# Getting Input keywords
wb1 = load_workbook(file_path)
sheet_input = wb1.get_sheet_by_name('Лист1')
# sheet_input.title

# Preparing new output file
now = datetime.now()
output_file = 'results_' + now.strftime("%Y-%m-%d_%H:%M:%S") + '.xlsx'
wb2 = openpyxl.Workbook()
sheetname = method + "_" + se
wb2.create_sheet("Top_keyword")
sheet = wb2.get_sheet_by_name('Top_keyword')

# Preparing output column names
sheet.cell(row=1, column=1).value = 'Keyword'
sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
sheet.cell(row=1, column=2).value = 'Response Message'
sheet.cell(row=1, column=2).font = openpyxl.styles.Font(bold=True)
sheet.cell(row=1, column=3).value = 'Found Results'
sheet.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
sheet.cell(row=1, column=4).value = 'Domains'
sheet.cell(row=1, column=4).font = openpyxl.styles.Font(bold=True)
sheet.cell(row=1, column=5).value = 'Full URLs'
sheet.cell(row=1, column=5).font = openpyxl.styles.Font(bold=True)

# Sending queries to web API and parsing data from responses
for i in range(int(start_cell), int(end_cell)):
    qr = sheet_input.cell(row=i, column=1).value.replace(" ", "%20")
    params = {
        'query': qr,  # string for get info
        'se': se ,  # string search engine
        'token': my_token,  # string personal token
    }
    api_url = "{host}/{method}?{params}".format(
        host=host,
        method=method,
        params=urlencode(params)
    )
    try:
        json_data = urlrequest.urlopen(api_url).read()
    except Exception as e0:
        print("API request error: {error}".format(error=e0))

    # Parsing JSON
    jsondata = json.loads(json_data)
    # Extracting data from JSON
    domains = ""
    urls = ""
    left_lines = jsondata['left_lines']
    status_msg = "%i" % (jsondata['status_code']) + ", " + jsondata['status_msg']
    results_found = ""

    if jsondata['status_code'] == 200:
        results_found = jsondata['result']['results']
        # Getting domains
        for item in jsondata['result']['top']:
            domains += item.get("domain") + ","
        # Getting URLs
        for item in jsondata['result']['top']:
            urls += item.get("url") + ","

    print(qr)
    print(status_msg)
    print(results_found)
    print(domains)
    print(urls)
    print("lines left: " + results_found)
    sheet.cell(row=i + 1, column=1).value = qr.replace("%20", " ")
    sheet.cell(row=i + 1, column=2).value = status_msg
    sheet.cell(row=i + 1, column=3).value = results_found
    sheet.cell(row=i + 1, column=4).value = domains
    sheet.cell(row=i + 1, column=5).value = urls

# Writing output to file
wb2.save(output_file)

# Opening window with results
root2 = Tk()
T = Text(root2, height=4, width=30)
B1 = Button(text="Ok", command = root2.destroy)
T.pack()
B1.pack()
T.insert(END, "Done")
T.insert(END, "\nОсталось запросов по токену:" + left_lines)
T.insert(END, "\nРезультат сохранен в файл: " + output_file)
mainloop()
"""